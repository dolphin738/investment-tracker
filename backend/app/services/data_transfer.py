"""数据导入导出服务（§4.2.17 · T05 · CSV/XLSX）。

职责：
- 导出 7 类（securities / securityTrades / cashFlows / cashBalances /
  securityPrices / assetSnapshots / navSeries）：CSV（UTF-8 BOM + `#` 注释行）/
  XLSX（openpyxl）。
- 模板 3 类（securityTrades / cashFlows / assetSnapshots）：表头 + 1 行示例。
- 导入预览（不落库）：解析 + 逐行校验（9 种错误码）+ 签发 10 分钟有效 token。
- 导入提交：持 token 在单事务内写入；事务提交后**仅调用 1 次** recalculateNavRange。

Decimal 一律字符串原样读写（不丢精度）。跨组合安全：以路径 portfolio_id 为准，
文件内 portfolioId 列忽略。
"""
from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation

import jwt
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.config import get_settings
from app.core.date_utils import today_app_tz
from app.core.enums import BusinessErrorCode
from app.core.exceptions import BusinessException
from app.models.enums import ExportType, ImportErrorCode, ImportType
from app.schemas_resp import ImportRowError
from app.models import (
    AssetSnapshot,
    CashBalance,
    CashFlow,
    DailyNav,
    DailyXirr,
    PortfolioSecurity,
    SecurityPrice,
    SecurityTrade,
)
from app.models.enums import CashFlowType, SecuritySide
from app.services.cashflow import CashflowService
from app.services.recalculation import RecalculationService
from app.services.security import compute_type
from app.services.snapshot import SnapshotService
from app.services.trade import TradeService

settings = get_settings()

MAX_FILE_BYTES = 5 * 1024 * 1024
MAX_ROWS = 10000
TOKEN_TTL_MIN = 10
ALLOWED_EXT = {".csv", ".xlsx", ".xls"}

EXPORT_TYPES = {e.value for e in ExportType}
IMPORT_TYPES = {e.value for e in ImportType}

# 导入字段规格：字段 -> 类型（date/decimal/enum/security/text）
_FIELD_KIND = {
    "securityTrades": {
        "date": "date",
        "securityCode": "security",
        "side": "enum",
        "quantity": "decimal",
        "costPrice": "decimal",
        "feeTotal": "decimal",
        "note": "text",
    },
    "cashFlows": {
        "date": "date",
        "type": "enum",
        "amount": "decimal",
        "note": "text",
    },
    "assetSnapshots": {
        "date": "date",
        "totalAsset": "decimal",
        "marketValue": "decimal",
        "cashBalance": "decimal",
        "note": "text",
    },
}
_REQUIRED = {
    "securityTrades": ["date", "securityCode", "side", "quantity", "costPrice"],
    "cashFlows": ["date", "type", "amount"],
    "assetSnapshots": ["date", "totalAsset"],
}
_ENUM_VALUES = {
    "side": {x.value for x in SecuritySide},
    "type": {x.value for x in CashFlowType},
}


# ── 文件解析 ──
def _read_sheet(content: bytes, ext: str) -> tuple[list[str], list[list[str]]]:
    """返回 (header, data_rows)。跳过空行与 `#` 开头行。"""
    if ext == ".csv":
        text = content.decode("utf-8-sig")
        raw = [r for r in csv.reader(io.StringIO(text))]
    else:  # xlsx / xls
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        raw = []
        for r in ws.iter_rows(values_only=True):
            raw.append(
                ["" if c is None else (_cell_to_str(c)) for c in r]
            )
    header: list[str] | None = None
    data: list[list[str]] = []
    for r in raw:
        if not any(str(c).strip() for c in r):
            continue
        if str(r[0]).lstrip().startswith("#"):
            continue
        if header is None:
            header = [str(c).strip() for c in r]
        else:
            data.append([str(c).strip() for c in r])
    if header is None:
        header = []
    return header, data


def _cell_to_str(c) -> str:
    if isinstance(c, datetime):
        return c.date().isoformat()
    if isinstance(c, date):
        return c.isoformat()
    return str(c)


def _ext_of(filename: str | None) -> str:
    if not filename:
        return ""
    return "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


# ── 校验 / 构造预览 ──
def _parse_date(s: str) -> date | None:
    s = (s or "").strip()
    if not s:
        return None
    try:
        return date.fromisoformat(s)
    except ValueError:
        # 兼容 2024/01/01
        try:
            return datetime.strptime(s, "%Y/%m/%d").date()
        except ValueError:
            return None


def _parse_decimal(s: str, max_scale: int = 2) -> Decimal | None:
    """解析十进制字符串；超过 max_scale 位小数视为精度非法返回 None。

    M4：数量/价格为 Numeric(18,6)，默认 2 位约束会误拒碎股/高精度报价；
    调用方按需传入 max_scale（如 quantity/price 用 6）。
    """
    s = (s or "").strip()
    if s == "":
        return None
    try:
        d = Decimal(s)
    except (InvalidOperation, ValueError):
        return None
    if d.as_tuple().exponent < -max_scale:
        return None
    return d


def _err(row, field, code: ImportErrorCode, message) -> ImportRowError:
    return ImportRowError(row=row, field=field, code=code, message=message)


def validate_and_build(
    type_: str,
    header: list[str],
    rows: list[list[str]],
    security_map: dict[str, str],
) -> tuple[list[dict], list[dict], list[dict], str | None]:
    """返回 (valid_rows, errors, sample, min_date_iso)。

    valid_rows：token 用的干净数据（security_id 已解析、date 为 ISO、decimal 为字符串）。
    errors：含全局错误（row=null）与行级错误。
    sample：前 10 行原始字符串样例（列->值）。
    """
    errors: list[ImportRowError] = []
    fields = _FIELD_KIND[type_]
    col_order = list(fields.keys())
    header_lower = {h.lower(): i for i, h in enumerate(header)}
    idx = {col: header_lower.get(col.lower()) for col in col_order}

    # 必需列缺失（全局错误）
    for col in _REQUIRED[type_]:
        if idx[col] is None:
            errors.append(_err(None, col, ImportErrorCode.MISSING_REQUIRED_COLUMN, f"缺少必需列：{col}"))

    # 文件过大
    if len(rows) > MAX_ROWS:
        errors.append(
            _err(None, None, ImportErrorCode.TOO_MANY_ROWS, f"行数超过上限 {MAX_ROWS}")
        )

    valid_rows: list[dict] = []
    seen_snap_dates: dict[str, int] = {}
    for i, r in enumerate(rows, start=1):
        row_errs: list[dict] = []
        parsed: dict = {}
        for col, kind in fields.items():
            pos = idx[col]
            raw = r[pos] if (pos is not None and pos < len(r)) else ""
            if kind == "date":
                d = _parse_date(raw)
                if d is None:
                    row_errs.append(_err(i, col, ImportErrorCode.INVALID_DATE_FORMAT, f"{col} 日期格式无效：{raw}"))
                else:
                    parsed[col] = d.isoformat()
            elif kind == "decimal":
                if raw == "" and col not in _REQUIRED[type_]:
                    parsed[col] = None  # 可选列可空
                    continue
                # 金额/费用限 2 位小数；数量/价格放宽到 6 位（Numeric(18,6)）
                max_scale = 6 if col in ("quantity", "costPrice") else 2
                d = _parse_decimal(raw, max_scale=max_scale)
                if d is None:
                    row_errs.append(_err(i, col, ImportErrorCode.INVALID_DECIMAL_PRECISION, f"{col} 数值无效（最多 {max_scale} 位小数）：{raw}"))
                else:
                    parsed[col] = str(d)
            elif kind == "enum":
                if raw not in _ENUM_VALUES[col]:
                    row_errs.append(_err(i, col, ImportErrorCode.INVALID_ENUM_VALUE, f"{col} 取值无效：{raw}"))
                else:
                    parsed[col] = raw
            elif kind == "security":
                sid = security_map.get(raw)
                if sid is None:
                    row_errs.append(_err(i, col, ImportErrorCode.SECURITY_NOT_FOUND, f"标的代码不存在于本组合：{raw}"))
                else:
                    parsed["security_id"] = sid
            else:  # text
                parsed[col] = raw

        # assetSnapshots 重复日期检测
        if type_ == "assetSnapshots" and "date" in parsed:
            d = parsed["date"]
            if d in seen_snap_dates:
                row_errs.append(_err(i, "date", ImportErrorCode.DUPLICATE_SNAPSHOT_DATE, f"同一日期出现多次：{d}"))
            else:
                seen_snap_dates[d] = i

        if not row_errs:
            valid_rows.append(parsed)
        else:
            errors.extend(row_errs)

    # sample：前 10 行原始字符串
    sample = []
    for r in rows[:10]:
        sample.append({col: (r[idx[col]] if (idx[col] is not None and idx[col] < len(r)) else "") for col in col_order})

    min_date = None
    if valid_rows:
        ds = [r["date"] for r in valid_rows if "date" in r]
        if ds:
            min_date = min(ds)

    return valid_rows, errors, sample, min_date


def make_token(type_: str, portfolio_id: str, valid_rows: list[dict], min_date: str | None) -> str:
    now = datetime.now(timezone.utc)
    payload = {
        "purpose": "dt_import",
        "type": type_,
        "portfolio_id": portfolio_id,
        "rows": valid_rows,
        "min_date": min_date,
        "iat": now,
        "exp": now + timedelta(minutes=TOKEN_TTL_MIN),
    }
    return jwt.encode(payload, settings.JWT_SECRET, algorithm=settings.JWT_ALGORITHM)


def decode_token(token: str) -> dict:
    try:
        return jwt.decode(token, settings.JWT_SECRET, algorithms=[settings.JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise BusinessException(
            code=BusinessErrorCode.VALIDATION_FAILED,
            message="导入令牌已过期，请重新预览",
            status_code=400,
        )
    except jwt.PyJWTError:
        raise BusinessException(
            code=BusinessErrorCode.VALIDATION_FAILED,
            message="导入令牌无效",
            status_code=400,
        )


# ── 导出 ──
async def build_export(
    type_: str, db: AsyncSession, portfolio_id: str
) -> tuple[list[str], list[list[str]]]:
    if type_ == "securities":
        secs = (
            await db.execute(
                select(PortfolioSecurity)
                .where(PortfolioSecurity.portfolio_id == portfolio_id)
                .options(selectinload(PortfolioSecurity.master))
            )
        ).scalars().all()
        cols = ["code", "name", "type", "currency"]
        rows = [
            [
                s.master.code if s.master else "",
                s.master.name or "" if s.master else "",
                compute_type(s).value,
                s.currency,
            ]
            for s in secs
        ]
        return cols, rows
    if type_ == "securityTrades":
        trades = (
            await db.execute(
                select(SecurityTrade)
                .where(SecurityTrade.portfolio_id == portfolio_id)
                .order_by(SecurityTrade.date, SecurityTrade.created_at)
            )
        ).scalars().all()
        secs = (
            await db.execute(
                select(PortfolioSecurity)
                .where(PortfolioSecurity.portfolio_id == portfolio_id)
                .options(selectinload(PortfolioSecurity.master))
            )
        ).scalars().all()
        code_map = {s.id: (s.master.code if s.master else "") for s in secs}
        cols = ["date", "securityCode", "side", "quantity", "costPrice", "feeTotal", "note"]
        rows = [
            [
                t.date.isoformat(),
                code_map.get(t.security_id, ""),
                t.side.value,
                str(t.quantity),
                str(t.cost_price),
                str(t.fee_total),
                t.note or "",
            ]
            for t in trades
        ]
        return cols, rows
    if type_ == "cashFlows":
        cfs = (
            await db.execute(
                select(CashFlow)
                .where(CashFlow.portfolio_id == portfolio_id)
                .order_by(CashFlow.date, CashFlow.created_at)
            )
        ).scalars().all()
        cols = ["date", "type", "amount", "note"]
        rows = [[c.date.isoformat(), c.type.value, str(c.amount), c.note or ""] for c in cfs]
        return cols, rows
    if type_ == "cashBalances":
        cbs = (
            await db.execute(
                select(CashBalance)
                .where(CashBalance.portfolio_id == portfolio_id)
                .order_by(CashBalance.as_of)
            )
        ).scalars().all()
        cols = ["asOf", "amount", "note"]
        rows = [[c.as_of.isoformat(), str(c.amount), c.note or ""] for c in cbs]
        return cols, rows
    if type_ == "securityPrices":
        ps = (
            await db.execute(
                select(SecurityPrice)
                .where(SecurityPrice.portfolio_id == portfolio_id)
                .order_by(SecurityPrice.as_of)
            )
        ).scalars().all()
        secs = (
            await db.execute(
                select(PortfolioSecurity)
                .where(PortfolioSecurity.portfolio_id == portfolio_id)
                .options(selectinload(PortfolioSecurity.master))
            )
        ).scalars().all()
        code_map = {s.id: (s.master.code if s.master else "") for s in secs}
        cols = ["asOf", "securityCode", "price"]
        rows = [[p.as_of.isoformat(), code_map.get(p.security_id, ""), str(p.price)] for p in ps]
        return cols, rows
    if type_ == "assetSnapshots":
        snaps = (
            await db.execute(
                select(AssetSnapshot)
                .where(AssetSnapshot.portfolio_id == portfolio_id)
                .order_by(AssetSnapshot.date)
            )
        ).scalars().all()
        cols = ["date", "totalAsset", "marketValue", "cashBalance", "source", "note"]
        rows = [
            [
                s.date.isoformat(),
                str(s.total_asset),
                str(s.market_value) if s.market_value is not None else "",
                str(s.cash_balance) if s.cash_balance is not None else "",
                s.source.value,
                s.note or "",
            ]
            for s in snaps
        ]
        return cols, rows
    if type_ == "navSeries":
        navs = (
            await db.execute(
                select(DailyNav)
                .where(DailyNav.portfolio_id == portfolio_id)
                .order_by(DailyNav.date)
            )
        ).scalars().all()
        xirrs = (
            await db.execute(select(DailyXirr).where(DailyXirr.portfolio_id == portfolio_id))
        ).scalars().all()
        snaps = (
            await db.execute(
                select(AssetSnapshot).where(AssetSnapshot.portfolio_id == portfolio_id)
            )
        ).scalars().all()
        xirr_map = {x.date: x.xirr_value for x in xirrs}
        snap_map = {s.date: s.total_asset for s in snaps}
        cols = ["date", "cumulativeNav", "yearlyNav", "shares", "totalAsset", "xirr"]
        rows = []
        for n in navs:
            xv = xirr_map.get(n.date)
            rows.append(
                [
                    n.date.isoformat(),
                    str(n.cumulative_nav),
                    str(n.year_nav),
                    str(n.shares),
                    str(snap_map[n.date]) if n.date in snap_map else "",
                    str(xv) if xv is not None else "",
                ]
            )
        return cols, rows
    raise BusinessException(
        code=BusinessErrorCode.VALIDATION_FAILED,
        message=f"未知导出类型：{type_}",
        status_code=400,
    )


def example_row(type_: str) -> list[str]:
    if type_ == "securityTrades":
        return ["2024-01-01", "600000", "BUY_SEC", "100", "10.50", "0", ""]
    if type_ == "cashFlows":
        return ["2024-01-01", "BUY", "100000.00", ""]
    if type_ == "assetSnapshots":
        return ["2024-01-01", "100000.00", "10000.00", "90000.00", ""]
    return []


def to_csv(columns: list[str], rows: list[list[str]], comment: str) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(columns)
    w.writerow(["# " + comment])
    for r in rows:
        w.writerow(r)
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def to_xlsx(columns: list[str], rows: list[list[str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(columns)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def safe_name(name: str) -> str:
    return re.sub(r"[^A-Za-z0-9_\-]", "_", name or "portfolio")[:40]


# ── 提交 ──
async def commit_import(
    db: AsyncSession,
    type_: str,
    portfolio_id: str,
    rows: list[dict],
    min_date: str | None,
) -> dict:
    inserted = updated = skipped = 0
    failed: list[ImportRowError] = []

    if type_ == "securityTrades":
        # 收口：证券买卖批量写入统一走 TradeService（卖出硬校验 + 构造），
        # 不再在导入层直接构造 ORM 模型，消除与 REST 写入的双真源（D10）。
        inserted += await TradeService(db).bulk_create(portfolio_id, rows)
    elif type_ == "cashFlows":
        # 收口：现金流水批量写入统一走 CashflowService（M1 校验 + 构造），
        # 不再在导入层直接构造 ORM 模型，消除与 REST 写入的双真源（D10）。
        await CashflowService(db).bulk_create(portfolio_id, rows)
        inserted += len(rows)
    elif type_ == "assetSnapshots":
        # 收口：总资产快照批量 upsert 统一走 SnapshotService（与 REST 单条 create 同源），
        # 不再在导入层直接调 AssetValuationService + 内联 select 计数（D10）。
        ins, upd = await SnapshotService(db).bulk_upsert(portfolio_id, rows)
        inserted += ins
        updated += upd

    await db.commit()

    recalculated = None
    if min_date:
        start_date = date.fromisoformat(min_date)
        if type_ == "assetSnapshots":
            # T5: only calculation layer cascade
            days = await RecalculationService(db).recalculateNavRange(
                portfolio_id, start_date
            )
        else:
            # T1-T4: snapshot layer rebuild + calculation layer cascade
            force = await RecalculationService(db).snapshot_dates_since(
                portfolio_id, start_date
            )
            days = await RecalculationService(db).recalculateRange(
                portfolio_id, start_date, force_dates=force
            )
        recalculated = {
            "fromDate": min_date,
            "toDate": today_app_tz().isoformat(),
            "recalculatedDays": days.affected_days,
        }
    return {
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "failed": failed,
        "recalculated": recalculated,
    }
